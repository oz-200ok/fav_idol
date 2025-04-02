from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl import load_workbook
from rest_framework.generics import (
    ListAPIView,
    ListCreateAPIView,
    RetrieveUpdateDestroyAPIView,
)
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from config.base_exception import ForbiddenException, ServiceUnavailableException
from config.permissions import IsAdminOrReadOnly
from Preferences.notification_service import NotificationService

from .models import Schedule
from .serializer import ScheduleSerializer
from .swagger_schema import (
    delete_response_schema,
    generate_swagger_response,
    update_create_response_schema,
)


class ScheduleListView(ListCreateAPIView):
    """
    일정 목록 조회 및 등록
    """

    queryset = Schedule.objects.all()
    serializer_class = ScheduleSerializer
    permission_classes = [IsAdminOrReadOnly]

    @swagger_auto_schema(request_body=ScheduleSerializer)
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request  # request 객체를 context에 추가
        return context

    def perform_create(self, serializer):
        group = serializer.validated_data.get("group")
        if not group:  # 그룹이 없으면 503 오류 반환
            raise ServiceUnavailableException(
                detail="선택된 그룹을 사용할 수 없습니다."
            )
        # 작성자를 자동으로 현재 사용자로 설정
        serializer.save(user=self.request.user)
        # 일정 생성 후 알림 발송
        NotificationService.notify_schedule_creation(serializer.instance)


class ScheduleDetailView(RetrieveUpdateDestroyAPIView):
    """
    특정 일정 상세 조회, 수정 및 삭제
    """

    queryset = Schedule.objects.all()
    serializer_class = ScheduleSerializer
    permission_classes = [IsAdminOrReadOnly]

    @swagger_auto_schema(
        request_body=ScheduleSerializer,
        responses={
            200: openapi.Response(
                description="일정 삭제 성공", schema=delete_response_schema
            ),
            404: openapi.Response(description="일정을 찾을 수 없음"),
        },
    )
    def delete(self, request, *args, **kwargs):
        instance = self.get_object()

        # 사용자가 일정의 작성자인지 확인
        if instance.user != request.user:
            raise ForbiddenException(detail="삭제 권한이 없습니다.")

        deleted_schedule_data = {
            "schedule.id": instance.id,
            "group_id": instance.group_id,
            "title": instance.title,
        }
        self.perform_destroy(instance)
        return Response({"data": deleted_schedule_data}, status=200)

    @swagger_auto_schema(
        request_body=ScheduleSerializer,
        responses={
            200: openapi.Response(
                description="일정 수정 성공", schema=update_create_response_schema
            ),
        },
    )
    def update(self, request, *args, **kwargs):
        # 일정 수정 처리 로직...
        updated_data = {
            "group_id": request.data.get("group_id"),
            "schedule_id": self.kwargs["pk"],
        }
        return Response({"data": updated_data}, status=200)

    @swagger_auto_schema(
        request_body=ScheduleSerializer,
        responses={
            201: openapi.Response(
                description="일정 등록 성공", schema=update_create_response_schema
            ),
        },
    )
    def create(self, request, *args, **kwargs):
        # 일정 등록 처리 로직...
        created_data = {
            "group_id": request.data.get("group_id"),
            "schedule_id": 123,  # 새로 생성된 ID를 반환
        }
        return Response({"data": created_data}, status=201)


class GroupScheduleListView(ListAPIView):
    """
    특정 그룹의 일정 필터링
    """

    serializer_class = ScheduleSerializer

    @swagger_auto_schema(
        responses=generate_swagger_response("특정 그룹 일정 조회", None),
    )
    def get_queryset(self):
        group_id = self.kwargs["group_id"]
        return Schedule.objects.filter(group_id=group_id).prefetch_related(
            "participating_members"
        )

    @swagger_auto_schema(
        responses=generate_swagger_response("그룹 일정 목록", None),
    )
    def list(self, request, *args, **kwargs):
        # 기존의 queryset 가져오기
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        # {"data": ...} 형식으로 리스폰스 반환
        return Response({"data": serializer.data})


class UserScheduleListView(ListAPIView):
    """
    본인이 작성한 일정 목록 조회
    """

    serializer_class = ScheduleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # 현재 사용자가 작성한 일정만 반환
        return Schedule.objects.filter(user=self.request.user).select_related("group")


class ExcelUploadview(ListCreateAPIView):
    """
    일정 등록 및 조회를 엑셀 파일을 업로드하여 진행합니다.
    """

    queryset = Schedule.objects.all()
    serializer_class = ScheduleSerializer
    permission_classes = [IsAdminOrReadOnly]
    parser_classes = (MultiPartParser, FormParser)

    def create(self, request, *args, **kwargs):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "엑셀 파일을 업로드 해주세요."}, status=400)
        try:
            # 엑셀 파일 로드
            workbook = load_workbook(file)
            sheet = workbook.active  # 첫번째 시트 사용

            # 엑셀 데이터 읽기
            schedules = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                (
                    group_id,
                    title,
                    description,
                    location,
                    start_time,
                    end_time,
                    participating_member_ids,
                ) = row
                schedule_data = {
                    "group": (
                        int(group_id) if group_id else None
                    ),  # 그룹 아이디 정수로 변환
                    "title": title,
                    "description": description,
                    "location": location,
                    "start_time": start_time,
                    "end_time": end_time,
                    "participating_member_ids": (
                        [int(x) for x in str(participating_member_ids).split(",")]
                        if participating_member_ids
                        else []
                    ),
                }
                # serializer를 통해 검증 및 저장
                serializer = self.get_serializer(data=schedule_data)
                serializer.is_valid(raise_exception=True)
                schedule = serializer.save(
                    user=request.user
                )  # 작성자를 현재 사용자로 설정

                # JSON 응답을 위해 Schedule 객체를 Serializer로 직렬화
                schedules.append(self.get_serializer(schedule).data)

            return Response({"data": schedules}, status=201)
        except Exception as e:
            return Response({"error": str(e)}, status=400)
